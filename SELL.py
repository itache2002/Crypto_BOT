import websocket
import json
from binance.client import Client
from talib import abstract
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime
import time 
import math
from binance.exceptions import BinanceAPIException, BinanceOrderException


class TradingBot:
    def __init__(self, api_key, api_secret, trading_symbol, leverage, risk_reward_ratio):
        self.api_key = api_key
        self.api_secret = api_secret
        self.trading_symbol =  trading_symbol
        self.leverage = leverage
        self.rrr = risk_reward_ratio
        self.client = Client(api_key, api_secret)
        self.df_final = None
        self.entry_price = 0.00
        self.exit_price = 0.00
        self.demo_acc = 10000
        self.tp = 0
        self.big_profit = 0
        self.stoploss = 0
        self.profit_loss = 0
        self.sec_big_profit = 0

    def set_initial_balance(self):
        start_balance = self.client.futures_account_balance()
        initial_balance = start_balance[5]['balance']
        print("================================")
        print('Initial balance:  {}'.format(initial_balance))
        print("================================")

    def set_leverage(self):
        change_leverage = self.client.futures_change_leverage(symbol=self.trading_symbol, leverage=self.leverage)
        print('Leverage set to: ', change_leverage['leverage'])

    def fetch_history_data(self):
        history_data = self.client.futures_klines(symbol=self.trading_symbol, interval=self.client.KLINE_INTERVAL_5MINUTE, limit=100)
        columns = ['T', 'O', 'H', 'L', 'C', 'V', 'KC', 'QV', 'NT', 'TBA', 'TBQ', 'E']
        self.df_final = pd.DataFrame(history_data, columns=columns)
        self.df_final= self.df_final.drop(self.df_final.tail(1).index)
        self.df_final= self.df_final.drop(['V', 'KC', 'QV', 'NT', 'TBA', 'TBQ', 'E'],axis= 1 )
        self.df_final['T'] = pd.to_datetime(self.df_final['T'], unit='ms')
        self.df_final['EMA'] = np.round(abstract.EMA(self.df_final['C'], timeperiod=5),2)
        print(self.df_final.tail(5))

    def place_order(self):
        exchange_info = self.client.get_exchange_info()
        symbol_info = next(item for item in exchange_info['symbols'] if item['symbol'] == self.trading_symbol)
        price_filter = next((filter for filter in symbol_info['filters'] if filter['filterType'] == 'PRICE_FILTER'), None)
        lot_size_filter = next((filter for filter in symbol_info['filters'] if filter['filterType'] == 'LOT_SIZE'), None)
        precision_for_price = int(-math.log10(float(price_filter['tickSize'])))
        precision_for_quantity = int(-math.log10(float(lot_size_filter['stepSize'])))

        check_position = self.client.futures_position_information()
        df = pd.DataFrame(check_position)
        position_amount = df.iloc[240]['positionAmt']
        TRADE_QUANTITY = (130 * self.leverage) / 42341
        half_TRADE_QUANTITY = TRADE_QUANTITY/2
        rounded_price = round(self.entry_price, precision_for_price)
        rounded_quantity = round(TRADE_QUANTITY, 3)
        round_half_quantity = round(half_TRADE_QUANTITY,precision_for_quantity)
        print(rounded_price)
        print(rounded_quantity)
        print(round_half_quantity)

        if float(position_amount) == 0:
            print("The entry price",self.entry_price)
            try:
                sell_limit_order =  self.client.futures_create_order(   symbol=self.trading_symbol, 
                                                                        side='SELL', 
                                                                        type='LIMIT', 
                                                                        timeInForce='GTC',
                                                                        price=rounded_price, 
                                                                        quantity=rounded_quantity)
                print('order placed')
                order_id=sell_limit_order['orderId']
                order_status = sell_limit_order['status']
                print(order_id)
                while order_status != 'FILLED':
                    time.sleep(10)
                    if order_status == 'FILLED':
                                time.sleep(1)
                                set_stop_loss = self.client.futures_create_order(symbol=self.trading_symbol,
                                                                                 side='BUY', 
                                                                                 type='STOP_MARKET', 
                                                                                 quantity=rounded_quantity, 
                                                                                 stopPrice=self.stoploss)
                                time.sleep(1)
                                set_take_profit = self.client.futures_create_order(symbol=self.trading_symbol, 
                                                                                   side='BUY', 
                                                                                   type='TAKE_PROFIT_MARKET', 
                                                                                   quantity=round_half_quantity, 
                                                                                   stopPrice=self.tp)
            except BinanceAPIException as e:
                    print(f"Binance API Exception: {e}")
                    print(f"Status Code: {e.status_code}")
                    print(f"Error Message: {e.message}")
            except BinanceOrderException as e:
                    # error handling goes here
                    print(e)

            

    def add_to_excel(self , timestamp ,entry_price, exit_price, stop_loss, take_profit, big_profit, pointes,sec_big_profit):
        try:
              wb = load_workbook('Sell_trades2.xlsx')
              sheet = wb.active
        except FileNotFoundError:
              wb = Workbook()
              sheet = wb.active
              sheet.append(['Entry Time', 'Entry Price', 'Exit Price','stop_loss','take_profit','big_profit','pointes','sec_big_profit'])

        for row in sheet.iter_rows(min_row=2, max_col=1):
            if row[0].value == timestamp:
                print("Duplicate entry found. Exiting without adding.")
                return
        sheet.append([timestamp, entry_price, exit_price, stop_loss, take_profit, big_profit,pointes,sec_big_profit])
        wb.save('Sell_trades2.xlsx')
        print("Entry added successfully.")

    def trailing_stoploss(self,current_price):
        points_covered = abs(current_price - self.entry_price)
        trailing_diff = abs(self.sec_big_profit - current_price)
        if points_covered > 200 and points_covered < 290 :
            self.stoploss = self.entry_price + 80
            if trailing_diff >= 40 and trailing_diff <=50 :
                self.stoploss =self.stoploss + 20
        if points_covered > 300 and points_covered < 390:
            if trailing_diff >= 10 and trailing_diff <=15:
                self.stoploss =self.stoploss + 10
        if points_covered > 400 and points_covered < 490 :
            if trailing_diff >= 5 and trailing_diff <= 10 :
                self.stoploss =self.stoploss + 10
        if points_covered > 500 :
            if trailing_diff >=5 and trailing_diff <=10:

                self.stoploss = self.stoploss + 5



    def run_trading_strategy_Sell(self, timestamp, last_open_price, last_close_price, last_high_price, last_low_price, last_EMA, is_Red,close_price):
      pointes = self.exit_price - self.entry_price
      closing_floor =int(close_price)
      if is_Red and last_low_price > last_EMA:
            print('##################################')
            print('SELL SIGNAL IS ON! Executing order')
            print('##################################')
            print("=========================================================")

            print("The time is :",timestamp)
            self.entry_price= last_close_price + 1
            print("The Entry price: ",self.entry_price)

            self.stoploss = last_high_price
            self.stoploss = int(self.stoploss)
            print("Calculated stop loss at:", self.stoploss)

            self.tp = self.entry_price - 60
            self.tp =  int(self.tp)
            print("Calculated Take profit at: ",self.tp)

            self.big_profit = self.tp - 100
            self.big_profit = int(self.big_profit)
            print("Calculated BIG Take profit at:", self.big_profit)

            self.sec_big_profit = self.big_profit - 100
            self.sec_big_profit = int(self.sec_big_profit)
            print("Calculated sec big profit profit at:",self.sec_big_profit)

            self.exit_price = self.stoploss
            print("Initial exit price set to stoploss:", self.exit_price)
            #placing the order
            self.place_order()


      if closing_floor == self.stoploss:
          self.exit_price = self.stoploss
          print("The stoploss is hit ")
          print(self.stoploss)
          print(close_price)
          self.add_to_excel(timestamp ,self.entry_price, self.exit_price , self.stoploss, self.tp, self.big_profit, pointes,self.sec_big_profit)


      if closing_floor == self.tp :
          self.exit_price = close_price
          print("Exit price updated to first take profit:", self.exit_price)
          self.add_to_excel(timestamp ,self.entry_price, self.exit_price , self.stoploss, self.tp, self.big_profit, pointes,self.sec_big_profit)

      if closing_floor == self.big_profit :
          self.exit_price = close_price
          print("Exit price updated to BIG take profit:", self.exit_price)
          self.add_to_excel(timestamp ,self.entry_price, self.exit_price , self.stoploss, self.tp, self.big_profit, pointes,self.sec_big_profit)

      if closing_floor == self.sec_big_profit:
          self.sec_big_profit =self.sec_big_profit - 100
          print("The new sec_big_profit is set:", self.sec_big_profit)
          self.trailing_stoploss(closing_floor)
          self.add_to_excel(timestamp ,self.entry_price, self.exit_price , self.stoploss, self.tp, self.big_profit, pointes,self.sec_big_profit)


        
    def Previous_Data(self,df, close_price):
        previous_data= df.iloc[-1]
        last_time= previous_data['T']
        last_close=float(previous_data['C'])
        last_open =float(previous_data['O'])
        last_low= float(previous_data['L'])
        last_high=float(previous_data['H'])
        last_EMA = previous_data['EMA']
        is_Red = last_close < last_open
        self.run_trading_strategy_Sell(last_time, last_open, last_close, last_high, last_low,last_EMA, is_Red, close_price)

        
    def Update_data(self, timestamp, open_price, high_price, low_price, close_price,candel_closing):
         History_data = self.df_final
         if candel_closing:
             new_data = pd.DataFrame({
                'T': [pd.to_datetime(timestamp , unit='ms')],
                'O': [open_price],
                'H': [high_price],
                'L': [low_price],
                'C': [close_price] })
             self.df_final= pd.concat([History_data, new_data], ignore_index=True) 
             self.df_final['EMA'] = np.round(abstract.EMA(self.df_final['C'] ,5),2)
         self.Previous_Data(self.df_final ,close_price)
        
    
    def on_message(self, ws, message):
        data = json.loads(message)
        candel = data['k']
        L_Time = candel['t']
        open_data = float(candel['o'])
        candel_closing = candel['x']
        high_data = float(candel['h'])
        low_data = float(candel['l'])
        closing_data = float(candel['c'])
        self.Update_data(L_Time, open_data, high_data, low_data, closing_data, candel_closing)

    def on_error(self, ws, error):
        print(f"WebSocket Error: {error}")

    def on_close(self, ws, close_status_code, close_msg):
        print("Closed")

    def on_open(self, ws):
        print("Connected")
        print('Receiving Data...')
        print("Waiting for Signal")

    def connect_websocket(self):
        websocket_url = f'wss://fstream3.binance.com/ws/{self.trading_symbol.lower()}@kline_5m'
        ws = websocket.WebSocketApp(websocket_url) 
        ws.on_open=self.on_open 
        ws.on_message=self.on_message 
        ws.on_error=self.on_error 
        ws.on_close=self.on_close
        ws.run_forever()

if __name__ == "__main__":
  
    api_key = 'api_key'
    api_secret = 'api_secret'
    # api_key = 'j3L0QNlc51TweyuFSQiqn9uSaE TUOYRdBDLRzpH9BV88avPryFoJ02c9QV31bxs1'
    # api_secret = '3tzp42W59aIdZ7LfRow89VAWP KOUv2QwaQMuNTfqEQ1t10GLG2cSuyOIbPbIFjc9'
    Trading_SYMBOL = 'BTCUSDT'

    # Initialize the trading bot
    trading_bot = TradingBot(api_key, api_secret, Trading_SYMBOL, 100, 2)

    # Set initial balance and leverage
    trading_bot.set_initial_balance()
    trading_bot.set_leverage()
    # # Fetch historical data
    # trading_bot.fetch_history_data()

   
    # Connect to WebSocket and start trading strategy
    # trading_bot.connect_websocket()
    # trading_bot.run_trading_strategy()



