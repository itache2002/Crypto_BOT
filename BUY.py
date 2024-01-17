import websocket
import json
from binance.client import Client
from talib import abstract
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime
import time 
from binance.exceptions import BinanceAPIException, BinanceOrderException
import math

class BUY():
    def __init__(self, api_key, api_secret, trading_symbol, leverage, risk_reward_ratio):
        self.api_key = api_key
        self.api_secret = api_secret
        self.trading_symbol = trading_symbol
        self.leverage = leverage
        self.rrr = risk_reward_ratio
        self.client = Client(api_key, api_secret)
        self.df_final = None
        self.entry_price = 0.00
        self.exit_price = 0.00
        self.tp = 0
        self.big_profit = 0
        self.stoploss = 0
        self.profit_loss = 0
        self.sec_big_profit = 0
        self.thi_big_profit = 0
        self.rounded_quantity = 0
        self.inter_stoploss_position = 0  
        self.sec_inter_stoploss_position =0
        self.thi_inter_stoploss_position = 0
        self.four_inter_stoploss_position = 0 
        self.five_inter_stoploss_position =0
        self.four_big_profit = 0
        self.five_big_profit =0

    def set_initial_balance(self):
        start_balance = self.client.futures_account_balance()
        initial_balance = start_balance[5]['balance']
        print("================================")
        print('Initial balance:  {}'.format(initial_balance))
        print("================================")

    def set_leverage(self):
        change_leverage = self.client.futures_change_leverage(symbol=self.trading_symbol, leverage=self.leverage)
        print('Leverage set to: ', change_leverage['leverage'])


    def fetch_history_data(self):
        history_data = self.client.futures_klines(symbol=self.trading_symbol, interval=self.client.KLINE_INTERVAL_15MINUTE, limit=100)
        columns = ['T', 'O', 'H', 'L', 'C', 'V', 'KC', 'QV', 'NT', 'TBA', 'TBQ', 'E']
        self.df_final = pd.DataFrame(history_data, columns=columns)
        self.df_final= self.df_final.drop(self.df_final.tail(1).index)
        self.df_final= self.df_final.drop(['V', 'KC', 'QV', 'NT', 'TBA', 'TBQ', 'E'],axis= 1 )
        self.df_final['T'] = pd.to_datetime(self.df_final['T'], unit='ms')
        self.df_final['EMA'] = np.round(abstract.EMA(self.df_final['C'], timeperiod=5),2)
        print(self.df_final.tail(5))
    def place_order(self):
        exchange_info = self.client.get_exchange_info()
        symbol_info = next(item for item in exchange_info['symbols'] if item['symbol'] == self.trading_symbol)

        price_filter = next((filter for filter in symbol_info['filters'] if filter['filterType'] == 'PRICE_FILTER'), None)
        lot_size_filter = next((filter for filter in symbol_info['filters'] if filter['filterType'] == 'LOT_SIZE'), None)

        precision_for_price = int(-math.log10(float(price_filter['tickSize'])))
        precision_for_quantity = int(-math.log10(float(lot_size_filter['stepSize'])))
        rounded_price = round(self.entry_price, precision_for_price)
        check_position = self.client.futures_position_information()
        df = pd.DataFrame(check_position)
        position_amount = df.iloc[240]['positionAmt']
        TRADE_QUANTITY = (32 * self.leverage) / rounded_price
        half_TRADE_QUANTITY = TRADE_QUANTITY/2
        self.rounded_quantity = round(TRADE_QUANTITY, 3)
        round_half_quantity = round(half_TRADE_QUANTITY,3)
        print(rounded_price)
        print(self.rounded_quantity)
        print(round_half_quantity)
        if float(position_amount) == 0:
            try:
                sell_limit_order =  self.client.futures_create_order(symbol=self.trading_symbol, 
                                                                     side='BUY', 
                                                                     type='LIMIT', 
                                                                     timeInForce='GTC', 
                                                                     price= rounded_price, 
                                                                     quantity=self.rounded_quantity)
                order_id=sell_limit_order['orderId']
                order_status = sell_limit_order['status']
                print(order_id)
                time.sleep(10)
                print(order_status)
                if order_status == 'FILLED' or order_status == 'NEW':
                    set_stop_loss = self.client.futures_create_order(symbol=self.trading_symbol, 
                                                                        side='SELL', 
                                                                        type='STOP_MARKET', 
                                                                        quantity=self.rounded_quantity, 
                                                                        stopPrice=self.stoploss)
                    

                    set_take_profit = self.client.futures_create_order(symbol=self.trading_symbol, 
                                                                        side='SELL', 
                                                                        type='TAKE_PROFIT_MARKET', 
                                                                        quantity=round_half_quantity, 
                                                                        stopPrice=self.tp)
                    
            except  BinanceAPIException as e:
                    print(f"Binance API Exception: {e}")
                    print(f"Status Code: {e.status_code}")
                    print(f"Error Message: {e.message}")
            except BinanceOrderException as e:
                    # error handling goes here
                    print(e)

    def add_to_excel(self , timestamp ,entry_price, exit_price, stop_loss, take_profit, big_profit, pointes, sec_big_profit,thi_big_profit):
        try:
              wb = load_workbook('BUY.xlsx')
              sheet = wb.active
        except FileNotFoundError:
              wb = Workbook()
              sheet = wb.active
              sheet.append(['Entry Time', 'Entry Price', 'Exit Price','stop_loss','take_profit','big_profit','pointes','sec_big_profit','thi_big_profit'])

        for row in sheet.iter_rows(min_row=2, max_col=7):
            existing_entry = [cell.value for cell in row]
            new_entry = [entry_price, exit_price, stop_loss, take_profit, big_profit, pointes,sec_big_profit,thi_big_profit]

            if existing_entry[1:] == new_entry:
                print("Duplicate entry found. Exiting without adding.")
                return

        sheet.append([timestamp, entry_price, exit_price,stop_loss,take_profit,big_profit,pointes,sec_big_profit,thi_big_profit])
        wb.save('BUY.xlsx')
        print("Entry added successfully.")


    def trailing_stoploss(self):
          
        # Cancel existing stop-loss order if it exists
          existing_stop_loss_orders = self.client.futures_get_open_orders(symbol=self.trading_symbol, side='SELL', type='STOP_MARKET')
          for order in existing_stop_loss_orders:
            order_id = order['orderId']
            self.client.futures_cancel_order(symbol=self.trading_symbol, orderId=order_id)
          # Create a new stop-loss order  
          set_stop_loss = self.client.futures_create_order(
            symbol=self.trading_symbol,
            side='SELL',
            type='STOP_MARKET',
            quantity=self.rounded_quantity,
            stopPrice=self.stoploss )
          
          print("Modified stop-loss order:", set_stop_loss)




    def run_trading_strategy_Sell(self, timestamp, last_open_price, last_close_price, last_high_price, last_low_price, last_EMA, is_Red,close_price):
      closing_floor =int(close_price)
      pointes = self.entry_price + closing_floor
      if not is_Red and last_high_price < last_EMA:
            print('##################################')
            print('BUY SIGNAL IS ON! Executing order')
            print('##################################')
            print("=========================================================")
            print("The time is :",timestamp)

            self.entry_price= last_close_price + 1
            print("The Entry price: ",self.entry_price)

            self.stoploss = last_low_price
            self.stoploss = int(self.stoploss)
            print("Calculated stop loss at:", self.stoploss)

            self.tp = self.entry_price + 50
            self.tp =  int(self.tp)
            print("Calculated Take profit at: ",self.tp)

            # frist big take profit  is 150 
            self.big_profit = self.tp + 100
            self.big_profit = int(self.big_profit)
            print("Calculated BIG Take profit at:", self.big_profit)

            # for intranl stoploss for 190 frist
            self.inter_stoploss_position =  self.big_profit + 40 
            self.inter_stoploss_position = int(self.inter_stoploss_position)

            # intrnal stoploss for 230 frist
            self.sec_inter_stoploss_position = self.inter_stoploss_position + 40
            self.sec_inter_stoploss_position = int(self.sec_inter_stoploss_position)

            #sec big take profit 260 
            self.sec_big_profit = self.big_profit + 110
            self.sec_big_profit = int(self.sec_big_profit)
            print("Calculated sec big profit profit at:",self.sec_big_profit)

            # intrnal stoploss for 300 for sec
            self.thi_inter_stoploss_position = self.sec_big_profit + 40
            self.thi_inter_stoploss_position =int(self.thi_inter_stoploss_position)

            # intrnal stoploss for 340  for  sec
            self.four_inter_stoploss_position = self.thi_inter_stoploss_position + 40
            self.four_inter_stoploss_position = int(self.four_inter_stoploss_position)
            
            # intrnal stoploss for 380  for  sec
            self.five_inter_stoploss_position =self.five_inter_stoploss_position + 40
            self.five_inter_stoploss_position = int(self.five_inter_stoploss_position)
        
            # third big take profit  400 
            self.thi_big_profit = self.five_inter_stoploss_position + 20
            self.thi_big_profit = int(self.thi_big_profit)
            print("Calculated 3rd big prifit profit at:",self.thi_big_profit)


            self.exit_price = self.stoploss
            print("Initial exit price set to stoploss:", self.exit_price)

            #placing the order
            self.place_order()

      if closing_floor == self.stoploss:
          self.exit_price = self.stoploss
          print("The stoploss is hit :",self.exit_price)
          self.add_to_excel(timestamp ,self.entry_price, self.exit_price , self.stoploss, self.tp, self.big_profit, pointes,self.sec_big_profit,self.thi_big_profit)

      if closing_floor == self.tp :
          self.exit_price = close_price
          print("Half price is sole at this point:", self.exit_price)
          print("Now the new stoploss is set as",self.stoploss)
          self.add_to_excel(timestamp ,self.entry_price, self.exit_price , self.stoploss, self.tp, self.big_profit, pointes,self.sec_big_profit,self.thi_big_profit)

      if closing_floor == self.big_profit :
          self.exit_price = close_price
          self.stoploss = self.entry_price
          print("The new stoploss is set as :",self.stoploss)
          self.add_to_excel(timestamp ,self.entry_price, self.exit_price , self.stoploss, self.tp, self.big_profit, pointes,self.sec_big_profit,self.thi_big_profit)
          #creating the a new order for every chang in stoploss 
          self.trailing_stoploss()

     #Setting the intral stoploss for 190
      if closing_floor ==self.inter_stoploss_position :
           self.stoploss = self.entry_price + 20
           #creating the a new order for every change in stoploss 
           self.trailing_stoploss()

     #setting the intrnal stoploss 230
      if closing_floor == self.sec_inter_stoploss_position:
           self.stoploss = self.stoploss + 20
           self.trailing_stoploss()

     #setting the intrnal stoploss 260
      if closing_floor == self.sec_big_profit:
          self.stoploss =  self.stoploss + 20
          print("The new stoploss is set as :",self.stoploss)
          self.add_to_excel(timestamp ,self.entry_price, self.exit_price , self.stoploss, self.tp, self.big_profit, pointes,self.sec_big_profit,self.thi_big_profit)
          #creating the a new order for every change in stoploss 
          self.trailing_stoploss()

      #setting the intrnal stoploss 300
      if closing_floor == self.thi_inter_stoploss_position:
           self.stoploss =self.stoploss + 20
           #creating the a new order for every chang in stoploss 
           self.trailing_stoploss()

      #setting the intrnal stoploss 340
      if closing_floor == self.four_inter_stoploss_position:
           self.stoploss =self.stoploss + 20
           #creating the a new order for every change in stoploss 
           self.trailing_stoploss()

      #setting the intrnal stoploss 380 
      if closing_floor == self.five_inter_stoploss_position:
           self.stoploss =self.stoploss + 20 
           #creating the a new order for every change in stoploss 
           self.trailing_stoploss()

      if closing_floor == self.thi_big_profit :
          self.thi_big_profit =  self.thi_big_profit + 20 
          self.stoploss =self.stoploss + 20
          print("The new stoploss is set as :",self.stoploss )
          print("The new 3rd big profit", self.thi_big_profit)
          self.add_to_excel(timestamp ,self.entry_price, self.exit_price , self.stoploss, self.tp, self.big_profit, pointes,self.sec_big_profit ,self.thi_big_profit)
          #creating the a new order for every change in stoploss 
          self.trailing_stoploss()
    
    
    def Previous_Data(self,df, close_price):
        previous_data= df.iloc[-1]
        last_time= previous_data['T']
        last_close=float(previous_data['C'])
        last_open =float(previous_data['O'])
        last_low= float(previous_data['L'])
        last_high=float(previous_data['H'])
        last_EMA = previous_data['EMA']
        is_Red = last_close < last_open
        self.run_trading_strategy_Sell(last_time, last_open, last_close, last_high, last_low,last_EMA, is_Red, close_price)


    def Update_data(self, timestamp, open_price, high_price, low_price, close_price,candel_closing):
         History_data = self.df_final
         if candel_closing:
             new_data = pd.DataFrame({
                'T': [pd.to_datetime(timestamp , unit='ms')],
                'O': [open_price],
                'H': [high_price],
                'L': [low_price],
                'C': [close_price] })
             self.df_final= pd.concat([History_data, new_data], ignore_index=True) 
             self.df_final['EMA'] = np.round(abstract.EMA(self.df_final['C'] ,5),2)
         self.Previous_Data(self.df_final ,close_price)


    def on_message(self, ws, message):
        data = json.loads(message)
        candel = data['k']
        L_Time = candel['t']
        open_data = float(candel['o'])
        candel_closing = candel['x']
        high_data = float(candel['h'])
        low_data = float(candel['l'])
        closing_data = float(candel['c'])
        self.Update_data(L_Time, open_data, high_data, low_data, closing_data, candel_closing)

    def on_error(self, ws, error):
        print(f"WebSocket Error: {error}")

    def on_close(self, ws, close_status_code, close_msg):
        print("Closed")

    def on_open(self, ws):
        print("Connected")
        print('Receiving Data...')
        print("Waiting for Signal")

    def connect_websocket(self):
        websocket_url = f'wss://fstream3.binance.com/ws/{self.trading_symbol.lower()}@kline_15m'
        ws = websocket.WebSocketApp(websocket_url) 
        ws.on_open=self.on_open 
        ws.on_message=self.on_message 
        ws.on_error=self.on_error 
        ws.on_close=self.on_close
        ws.run_forever()


if __name__ == "__main__":
    api_key = 'BwtkmrnTu9U2r1gjdy5gwtv3s8s82QR0kEt130MBNWLDMcImeJHU6Af8fyYpF7AN'
    api_secret = '9rJYUHzOrXnYdGY0dnsZwvcnmOVtrN1p8cHZwOtVeaAqRniY23A89Y8oMBRzeacF'
    Trading_SYMBOL = 'BTCUSDT'

    buy = BUY(api_key, api_secret, Trading_SYMBOL, 100, 2)

    
    buy.set_initial_balance()
    buy.set_leverage()
    buy.fetch_history_data()
    buy.connect_websocket()

    